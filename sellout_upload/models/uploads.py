# -*- coding: utf-8 -*-

from odoo import models, fields, api , tools
from odoo.exceptions import UserError
import datetime 

import tempfile
import binascii
import openpyxl
import io

import logging
_logger = logging.getLogger(__name__)


xlscolumns = [ 'DISTRIBUTOR','KODE_CTM','NAMA_CTM','ALAMAT','KELOMPOK_CUSTOMER',
               'KODE_WILAYAH','NAMA_WILAYAH',
               'KODE_SALES','NAMA_SALES',
               'TANGGAL','NO_FAKTUR',
               'KODE_BRG','NAMA_BRG',
               'QTY_JUAL','QTY_BONUS','HARGA_JUAL','RP_DISCOUNT','PCT_DISCOUNT']   



class SelloutHNAPricelist(models.Model):
    _name = 'sellout.hnapricelist'
    _description = 'Sellout HNA Pricelist'

    barang_id   = fields.Many2one(comodel_name='product.product', string='Link Barang')
    date_start  = fields.Date("Start")
    date_end    = fields.Date("End")
    price       = fields.Float("Price")


class SelloutUploadDetail(models.Model):
    _name = 'sellout.upload.detail'
    _description = 'Temp table upload'

    upload_id   = fields.Many2one('sellout.upload',string='upload id')
    company_id  = fields.Integer()
    kode_ctm    = fields.Char(string='Kode Pelanggan')
    nama_ctm    = fields.Char(string='Nama Pelanggan')
    kelom_cust  = fields.Char(string='Kelompok Pelanggan')
    alamat      = fields.Char(string='Alamat Pelanggan')
    kode_wilay  = fields.Char(string='Kode Wilayah')
    nama_wilay  = fields.Char(string='Nama Wilayah')
    kode_sls    = fields.Char(string='Kode Salesman')
    nama_sls    = fields.Char(string='Nama Salesman')
    tanggal     = fields.Date(string='Tanggal')
    no_faktur   = fields.Char(string='Reff')
    kode_brg    = fields.Char(string='Kode Barang')
    nama_brg    = fields.Char(string='Nama Barang')
    qty_jual    = fields.Float(string='QtyJual')
    qty_extra   = fields.Float(string='QtyExtra')
    a_rp_jl     = fields.Float(string='@Jual')
    a_rp_disc   = fields.Float(string='@Disc')
    a_rp_hna    = fields.Float(string='@Hna')
    rp_jual     = fields.Float(string='RpJual')
    rp_disc     = fields.Float(string='RpDisc')
    rp_hna      = fields.Float(string='RpHna')
    customer_id = fields.Integer()
    wilayah_id  = fields.Integer()
    barang_id   = fields.Integer()
    user_id     = fields.Integer()
    salesman_id = fields.Integer()
    klp_cust_id = fields.Integer()


class SelloutUpload(models.Model):
    _name = 'sellout.upload'
    _description = 'Sellout Upload'
    _check_company_auto = True

    name = fields.Char(string='Judul', required=True)
    description = fields.Text(string='Keterangan')  

    import_file = fields.Binary(string='Import File')
    filename    = fields.Char(string='File Name')
    template    = fields.Many2one('sellout.template',string='Template')
    company_id = fields.Many2one(
        comodel_name='res.company',
        index=True,
        default=lambda self: self.env.company)
    state       = fields.Selection([
        ('draft' , 'Draft'),
        ('upload', 'Upload'),
        ('post' , 'Posted'),
        ('cancel', 'Canceled'),
    ], string='Status' , readonly=True , 
        copy=False , index=True , default='draft')
    date_start     = fields.Date(string='Tgl.Mulai',invisible='1')        
    date_end       = fields.Date(string='Tgl.Akhir',invisible='1')        
    # detail 
   
   
    faktur_lines   = fields.One2many(comodel_name='sellout.upload.detail' , inverse_name='upload_id' , string='Fakturs')    
    customer_lines = fields.One2many(comodel_name='sellout.upload.detail.customer',inverse_name='upload_id',string='Pelanggan')
    barang_lines   = fields.One2many(comodel_name='sellout.upload.detail.barang' ,inverse_name='upload_id',string='Barang')
    wilayah_lines  = fields.One2many(comodel_name='sellout.upload.detail.wilayah',inverse_name='upload_id',string='Wilayah')
    so_lines       = fields.One2many(comodel_name='sale.order',inverse_name='upload_id',string='sale.order')
    salesman_lines = fields.One2many(comodel_name='sellout.upload.detail.salesman',inverse_name='upload_id',string='Salesman')   


    def unlink(self):
        if self.state in ('post','done'):
            raise ValidationError("Anda tidak bisa menghapus record ini ")
        
        ## hapus data 
        soline = self.env["sale.order.line"].search([('upload_id','=', self.id)])
        soline.unlink()
        so = self.env["sale.order"].search([('upload_id','=', self.id)])
        so.unlink()

        temp = self.env["sellout.upload.detail"].search([('upload_id','=', self.id)])
        temp.unlink()

        return super(SelloutUpload, self).unlink()


    def action_download(self):
        output = io.ByteIO()
        
        response = request.make_response(
                    None,
                    headers=[
                        ('Content-Type', 'application/vnd.ms-excel'),
                        ('Content-Disposition', content_disposition('UploadTemplate' + '.xlsx'))
                    ]
        )        


    def action_upload(self):
        nmax_record = 8000


        if self.company_id == 0 or self.import_file == "" or self.name == ""  :
            raise UserError("Data Upload belum lengkap , lengkapi Perusahaan , Nama file dan file ")
            return

        try:
            fp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            fp.write(binascii.a2b_base64(self.import_file))
            fp.seek(0)
            fp.close
        except FileNotFoundError:
            raise UserError('No such file or directory found. \n%s.' % self.file_name)            
        except:
            raise UserError("Invalid File!")

        workbook = openpyxl.load_workbook(fp.name)
        sheet    = workbook.active

        if sheet.max_row == 0:
            raise UserError("Empty File")
            return

        if sheet.max_row > nmax_record:
            raise UserError("Record Limit Exceded " )
            return             



        # delete dulu kalau sudah pernah di upload 
        # delete yang sale.order , harusnya sale.order.line nya juga kehapus 
        if not ( self.state=='' or self.state=='draft' or self.state=='done' ):
            soline = self.env["sale.order.line"].search([('upload_id','=', self.id)])
            soline.unlink()
            so = self.env["sale.order"].search([('upload_id','=', self.id)])
            so.unlink()

        temp = self.env["sellout.upload.detail"].search([('upload_id','=', self.id)])
        temp.unlink()
   
        ret_upload   = self.__upload_load_xls(sheet)
        if ret_upload==0:
            ret_barang   = self.__upload_barang()
            ret_salesman = self.__upload_salesman()
            ret_wilayah  = self.__upload_wilayah()
            ret_customer = self.__upload_customer()

            ## Mandatory barang tidak ada belum di posting 
            if ret_barang == 1:
                notification = {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': 'Warning Mandatory Link Table Empty',
                        'type': 'warning',
                        'message': 'Link Barang masih ada yang kosong',
                        'sticky': False,
                    }
                }
                return notification

            self.env.cr.commit()
            ret_post      = self.__post()

            if ret_salesman==0 and ret_wilayah==0 and ret_customer==0:
                notification = {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': 'Complete',
                        'type': 'success',
                        'message': 'File berhasil di upload',
                        'sticky': False,
                    }
                }
                self.write({'state' : 'post' })
                return notification
            endif

            self.env.cr.commit()
            ret_post      = self.__post()
            if ret_post ==0:
                notification = {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': 'INCOMPLETE POST',
                        'type': 'warning',
                        'message': 'Post incomplete',
                        'sticky': False,
                    }
                }
                self.write({'state' : 'upload' })
                return notification
            else:                
                notification = {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': 'INCOMPLETE UPLOAD',
                        'type': 'warning',
                        'message': 'Link masih ada yang kosong',
                        'sticky': False,
                    }
                }
                self.write({'state' : 'upload' })
                return notification


    def __upload_load_xls(self,sheet):

        xdate_start = None 
        xdate_end   = None 

        try:
            for row in range(2,sheet.max_row):

                xkode_ctm       = str(sheet.cell(row,2).value).rstrip()
                xnama_ctm       = sheet.cell(row,3).value
                xalamat         = sheet.cell(row,4).value
                xkelom_cust     = sheet.cell(row,5).value
                xkode_wilay     = str(sheet.cell(row,6).value).rstrip()
                xnama_wilay     = sheet.cell(row,7).value
                xkode_sls       = str(sheet.cell(row,8).value).rstrip()
                xnama_sls       = sheet.cell(row,9).value
                xtanggal        = sheet.cell(row,10).value
                if xtanggal in (datetime, datetime.date, datetime.datetime, datetime.time):
                    dt = datetime.datetime(1899,12,30) + datetime.timedelta(days=xtanggal)
                else:
                    dt = xtanggal 


                xno_faktur      = sheet.cell(row,11).value,
                xkode_brg       = sheet.cell(row,12).value.rstrip()
                xnama_brg       = sheet.cell(row,13).value    


                temp = self.env["sellout.upload.detail"].create({
                    "upload_id"     : self.id , 
                    "company_id"    : self.company_id.id ,
                    "kode_ctm"      : xkode_ctm,
                    "nama_ctm"      : xnama_ctm,
                    "alamat"        : xalamat,
                    "kelom_cust"    : xkelom_cust,
                    "kode_wilay"    : xkode_wilay,
                    "nama_wilay"    : xnama_wilay,
                    "kode_sls"      : xkode_sls,            
                    "nama_sls"      : xnama_sls,
                    "tanggal"       : dt,
                    "no_faktur"     : sheet.cell(row,11).value,
                    "kode_brg"      : xkode_brg,            
                    "nama_brg"      : xnama_brg,
                    "qty_jual"      : sheet.cell(row,14).value,
                    "qty_extra"     : sheet.cell(row,15).value,
                    "a_rp_jl"       : sheet.cell(row,16).value,
                    "a_rp_disc"     : sheet.cell(row,17).value,            
                    "a_rp_hna"      : 0,
                    "rp_jual"       : 0,
                    "rp_disc"       : 0,
                    "rp_hna"        : 0,
                    "customer_id"   : 0,
                    "salesman_id"   : 0,
                    "barang_id"     : 0,
                    "klp_cust_id"   : 0,
                    "wilayah_id"    : 0,
                })  
                if xdate_start is None:
                    xdate_start = dt 
                elif dt < xdate_start:
                    xdate_start = dt
                if xdate_end is None:
                    xdate_end = dt 
                elif dt > xdate_end:
                    xdate_end = dt

            self.date_start = xdate_start
            self.date_end   = xdate_end
        except Exception as e:
            raise UserError(e)
            return 1


        # check tanggal 
        cek_tanggal = 0
        tuploads = self.env["sellout.upload"].search([('company_id','=',self.company_id.id)])            
        for tupload in tuploads:
            if not tupload.id == self.id:
                if self.date_start >= tupload.date_start and self.date_start <= tupload.date_end:                
                    cek_tanggal = 1
                    raise UserError('Periode Data pernah dimasukkan')


                if self.date_end >= tupload.date_start and self.date_end <= tupload.date_end:
                    cek_tanggal = 1
                    raise UserError('Periode Data pernah dimasukkan')




        ## search date 
        ## self.date_start between date_start or date_end
        ## self.date_end   between date_start or date_end
        ## company_id = self.company_id.id and (
        # ( self.date_start >= date_start and self.date_start <= date_end )
        # ( self.date_end >= date_start and self.date_end <= date_end ))  
           
        #collab = self.env['sellout.upload'].search_count(['&',
        #    ('company_id','=',self.company_id.id),'|','&',
        #    ('date_start','>=',self.date_start),('date_end','<=',self.date_start),'&',
        #    ('date_start','>=',self.date_end),('date_end','<=',self.date_end),      
        #    ])
        #if collab>=0:
        #    raise UserError('Periode Data sudah pernah dimasukkan ')
        return 0



    def __upload_barang(self):
        xret = 0 
        barangs = self.env['sellout.upload.detail']._read_group( [
                    ('upload_id','=',self.id)
                ], ['company_id','kode_brg','nama_brg'],
                    ['__count'] )
        for barang in barangs:
            kode_brg = barang[1]
            nama_brg = barang[2]

            ketemu = self.env["sellout.link.barang"].search_count([
                    ('company_id','=',self.company_id.id),
                    ('code', '=', kode_brg ),
                    ('name', '=', nama_brg )         
            ])
            if ketemu==0:
                xret   = 1
                self.env["sellout.link.barang"].create({
                    'company_id' : self.company_id.id,
                    'code'       : kode_brg,
                    'name'       : nama_brg,
                    'barang_id'  : 0                    
                })
            else:
                kode = self.env["sellout.link.barang"].search([
                    ('company_id','=',self.company_id.id),
                    ('code', '=', kode_brg ),
                    ('name', '=', nama_brg )         
                ],limit=1)

                barang = self.env['sellout.upload.detail'].search([
                    ('upload_id','=' ,self.id),
                    ('company_id','=',self.company_id.id),
                    ('kode_brg', '=', kode_brg ),
                ])
                if kode.barang_id == 0:
                   xret = 1

                barang.write({
                    'barang_id' : kode.barang_id 
                })
                if kode.barang_id.id > 0:
                    ## update HHA 
                    hnas = self.env["sellout.hnapricelist"].search([('barang_id','=',kode.barang_id.id)],order='date_start')
                    for hna in hnas:
                        temps = self.env["sellout.upload.detail"].search([
                        ('upload_id','=',self.id),
                        ('barang_id','=',kode.barang_id.id),
                        ('tanggal','>=',hna.date_start),('tanggal','<=',hna.date_end)
                        ])
                        temps.write({
                            'a_rp_hna' : hna.price,
                        })

        return xret


    #salesman (NON-MANDATORY)
    def __upload_salesman(self):
        salesmans = self.env['sellout.upload.detail']._read_group( [
                  ('upload_id','=',self.id)
                ], ['company_id','kode_sls','nama_sls'],
                   ['__count'] )
        for salesman in salesmans:
            kode_sls = salesman[1]
            nama_sls = salesman[2]

            ketemu = self.env["sellout.link.salesman"].search_count([
                   ('company_id','=',self.company_id.id),
                   ('code', '=', kode_sls ),
                   ('name', '=', nama_sls )         
            ])
            if ketemu==0:
                ## version 2.0 
                ## auto create salesman master 
                nvalid = 1
                msalesman = self.env["sellout.salesman"].create({
                    'company_id' : self.company_id.id,
                    'kode_sls'   : kode_sls,
                    'nama_sls'   : nama_sls,               
                })
                
                self.env["sellout.link.salesman"].create({
                    'company_id' : self.company_id.id,
                    'code'       : kode_sls,
                    'name'       : nama_sls,
                    'salesman_id': msalesman.id ,
                })

            else:
                msalesman = self.env["sellout.link.salesman"].search([
                    ('company_id','=',self.company_id.id),
                    ('code', '=', kode_sls ),
                    ('name', '=', nama_sls )         
                ])

                salesman = self.env['sellout.upload.detail'].search([
                    ('upload_id','=' ,self.id),
                    ('company_id','=',self.company_id.id),
                    ('kode_sls', '=', kode_sls ),
                ])                
                salesman.write({
                    'salesman_id' : msalesman.salesman_id
                })   
        return 0

    #wilayah
    def __upload_wilayah(self):
        # INCOMPLETE code 
        
        external = self.env.ref('sellout_base.ct_00_000')

        xret = 0
        wilayahs = self.env['sellout.upload.detail']._read_group( [
                  ('upload_id','=',self.id)
                ], ['company_id','kode_wilay','nama_wilay'],
                   ['__count'] )
        for wilayah in wilayahs:
            kode_wilay = wilayah[1]
            nama_wilay = wilayah[2]

            ketemu = self.env["sellout.link.wilayah"].search_count([
                   ('company_id','=',self.company_id.id),
                   ('code', '=', kode_wilay ),
                   ('name', '=', nama_wilay )         
            ])
            if ketemu==0:
                xret   = 1
                self.env["sellout.link.wilayah"].create({
                    'company_id' : self.company_id.id,
                    'code'       : kode_wilay,
                    'name'       : nama_wilay,
                    'village_id' : 0,
                    'sub_district_id': 0,
                    'district_id': external.id,
                    'state_id'   : 0,
                })
            else:
                kode = self.env["sellout.link.wilayah"].search([
                    ('company_id','=',self.company_id.id),
                    ('code', '=', kode_wilay ),
                    ('name', '=', nama_wilay )         
                ])
                wilayah = self.env['sellout.upload.detail'].search([
                    ('upload_id',  '=',self.id),
                    ('company_id', '=',self.company_id.id),
                    ('kode_wilay', '=', kode_wilay ),
                ])                                   
                wilayah.write({
                    'wilayah_id': kode.district_id,
                })
        return xret 

    def __upload_customer(self):
        # MANDATORY tapi tidak perlu lengkap wilayah & kategory pelanggan 
        wilayah_external  = self.env.ref('sellout_base.ct_00_000')
        kelompok_external = self.env.ref('sellout_base._categ_0_0')
        xret              = 0

        customers = self.env['sellout.upload.detail']._read_group( [
                  ('upload_id','=',self.id)
                ], ['company_id','kode_ctm','nama_ctm','alamat','kelom_cust','kode_wilay','nama_wilay'],
                   ['__count'] )
        for customer in customers:
            kode_ctm   = customer[1]
            nama_ctm   = customer[2]
            alamat     = customer[3]
            kelom_cust = customer[4].rstrip()
            kode_wilay = customer[5].rstrip()
            nama_wilay = customer[6]
            customer_id = fields.Many2one(comodel_name='res.partner', string='Customer' )
            wilayah_id  = wilayah_external.id
            kelompok_id = kelompok_external.id

            ketemu = self.env["sellout.link.customer"].search_count([
                   ('company_id','=',self.company_id.id),
                   ('code', '=', kode_ctm ),
                   ('name', '=', nama_ctm )         
            ])
            if ketemu==0:
                nvalid = 1
                ketemu = self.env["sellout.link.customer.kelompok"].search_count([
                   ('company_id','=',self.company_id.id),
                   ('code', '=', kelom_cust ),
                ])

                if ketemu==0:
                    kelompok = self.env["sellout.link.customer.kelompok"].create({
                        'company_id' : self.company_id.id,
                        'code'       : kelom_cust ,
                        'klp_cust_id': kelompok_id,
                    })
                else:
                    kelompok = self.env["sellout.link.customer.kelompok"].search([
                    ('company_id','=',self.company_id.id),
                    ('code', '=', kelom_cust ),
                    ],limit=1)
                    if not ( kelompok.id == 0 or kelompok.id == kelompok_external.id ):
                        kelompok_id = kelompok.klp_cust_id.id
                
                ketemu = self.env["sellout.link.wilayah"].search_count([
                        ('company_id','=',self.company_id.id),
                        ('code'      ,'=',kode_wilay ),
                ])

                if ketemu==0:
                    wilayah = self.env["sellout.link.wilayah"].create({
                        'company_id' : self.company_id.id,
                        'code'       : kode_wilay ,
                        'district_id': wilayah_id,
                    })
                else:
                    wilayah = self.env["sellout.link.wilayah"].search([
                        ('company_id','=',self.company_id.id),
                        ('code', '=', kode_wilay ),
                        ],limit=1)
                    if not ( wilayah.id==0 or wilayah.id == wilayah_external.id ):
                        wilayah_id = wilayah.district_id.id

                # create INCOMPLETE customer master             

                mcustomer = self.env["res.partner"].create({
                    'company_id' : self.company_id.id,
                    'dgExType'   : kelom_cust,
                    'dgExNama'   : nama_ctm,
                    'dgExAlamat' : alamat,
                    'dgExWilayah': kode_wilay,
                    'dgCompanyId': self.company_id.id,
                    'dgExreff'   : kode_ctm,
                    'ref'        : kode_ctm,
                    'name'       : nama_ctm,
                    'street'     : alamat,
                    'district_id': wilayah_id,
                    'industry_id': kelompok_id,
                })          

    
                self.env["sellout.link.customer"].create({
                    'company_id' : self.company_id.id,
                    'code'       : kode_ctm,
                    'name'       : nama_ctm,
                    'alamat'     : alamat,
                    'kelom_cust' : kelom_cust,
                    'kode_wilay' : kode_wilay,
                    'nama_wilay' : nama_wilay,
                    'customer_id': mcustomer.id                   
                })
            else:

                kelompok = self.env["sellout.link.customer.kelompok"].search([
                    ('company_id','=',self.company_id.id),
                    ('code', '=', kelom_cust ),
                    ],limit=1)              
                if not ( kelompok.id ==0 or kelompok.id == kelompok_external.id ):
                    kelompok_id = kelompok.id

                wilayah = self.env["sellout.link.wilayah"].search([
                    ('company_id','=',self.company_id.id),
                    ('code', '=', kode_wilay ),
                    ],limit=1)
                if not ( wilayah.id==0 or wilayah.id == wilayah_external.id ):
                    wilayah_id = wilayah.district_id

                kode = self.env["sellout.link.customer"].search([
                    ('company_id','=',self.company_id.id),
                    ('code', '=', kode_ctm ),
                    ('name', '=', nama_ctm )         
                ])
                temps = self.env['sellout.upload.detail'].search([
                    ('upload_id','=' ,self.id),
                    ('company_id','=',self.company_id.id),
                    ('kode_ctm', '=', kode_ctm ),
                ])          
                temps.write({
                    'customer_id' : kode.customer_id 
                })

                customer = self.env["res.partner"].search([('id','=',kode.customer_id.id)],limit=1)

                if not ( kelompok_id == 0 or kelompok_id == kelompok_external.id ):
                    if customer.industry_id == 0 or customer.industry_id == kelompok_external.id: 
                        customer.write({
                            'industry_id' : kelompok_id
                        })

                if not wilayah_id == 0 or wilayah_id == wilayah_external.id : 
                    if customer.district_id == 0 or customer.district_id == wilayah_external.id : 
                        customer.write({
                            'district_id' : wilayah_id
                        })

                ## update data
                #          
        return xret  

    def __post(self):
        tempuploads = self.env["sellout.upload.detail"].search([('upload_id','=',self.id)], order='tanggal,no_faktur,kode_ctm')
        tanggal   = datetime.date.today()
        no_faktur = ''
        kode_ctm  = ''
        for t_upload in tempuploads:
            ## print(t_upload.tanggal,t_upload.no_faktur,t_upload.company_id,t_upload.customer_id,self.id,t_upload.salesman_id)
            ## 
            if not ( t_upload.customer_id == 0 or t_upload.barang_id == 0 or t_upload.salesman_id == 0 ):                
                if t_upload.tanggal != tanggal or t_upload.no_faktur != no_faktur or t_upload.kode_ctm != kode_ctm:
                    tanggal   = t_upload.tanggal 
                    no_faktur = t_upload.no_faktur
                    kode_ctm  = t_upload.kode_ctm

##                    print(t_upload)
##                    print("data : " , t_upload.no_faktur, t_upload.tanggal , t_upload.kode_ctm , t_upload.customer_id , t_upload.kode_sls , t_upload.salesman_id )
                    so = self.env["sale.order"].create({
                        'name'          : t_upload.no_faktur , 
                        'company_id'    : t_upload.company_id,
                        'date_order'    : t_upload.tanggal , 
                        'name'          : t_upload.no_faktur,
                        'partner_id'    : t_upload.customer_id,
                        'client_order_ref' : t_upload.no_faktur,
                        'origin'        : self.name,
                        'upload_id'     : self.id,
                        'user_id'        : t_upload.user_id,
                        'salesman_id'    : t_upload.salesman_id,
                        'currency_id'   : self.company_id.currency_id,
                    })
                    print("save order : " , t_upload.no_faktur, t_upload.tanggal , t_upload.kode_ctm )
                ## sale order line     
                if t_upload.qty_jual != 0:
                    so_line = self.env["sale.order.line"].create({
                        "order_id"        : so.id,
                        "product_id"      : t_upload.barang_id,
                        'price_unit'      : t_upload.a_rp_jl,
                        'product_uom_qty' : t_upload.qty_jual,
                        'a_rp_hna'        : t_upload.a_rp_hna,
                        'rp_hna'          : t_upload.a_rp_hna * t_upload.qty_jual,
                        'qty_extra'       : 0,
                        'upload_detail_id': t_upload.id,
                    })
                if t_upload.qty_extra != 0:
                    so_line = self.env["sale.order.line"].create({
                        "order_id"        : so.id,
                        "product_id"      : t_upload.barang_id,
                        'price_unit'      : 0,
                        'product_uom_qty' : t_upload.qty_extra,
                        'a_rp_hna'        : 0,
                        'rp_hna'          : 0,
                        'qty_extra'       : t_upload.qty_extra,
                        'upload_detail_id': t_upload.id,                        
                    })

    def action_confirm(self):
        # test 
        adayangkosong = self.env["sellout.upload.detail"].search_count([
            '&',('upload_id','=',self.id), '|',
            ('customer_id','=',0),('salesman_id','=',0),('barang_id','=',0),('wilayah_id','=',0)
        ])
        print(adayangkosong)
        if adayangkosong >0:
            raise UserError('Data Pelanggan / Salesman / Wilayah / Barang ada yang belum lengkap')            
            return

        self.env.cr.commit()
        print(" Execute query upload ")
        # tempupload = self.env["sellout.upload.detail"].search([('upload_id)','=',self.id)], order='tanggal,no_faktur,kode_ctm')
        # tempupload = self.env["sellout.upload.detail"].search([('upload_id','=',self.id)])
        # tempuploads = self.env["sellout.upload.detail"].search([('upload_id)','=',self.id)], order='tanggal,no_faktur,kode_ctm')



    def action_draft(self):
        if self.state != 'done':
            so = self.env['sale.order'].search([('upload_id','=',self.id)])
            so.unlink()
            self.state = 'draft'
        

    def action_done(self):
        if self.state == 'done':
           self.env.cr.execute("update state = 'sale'   , invoice_status = 'to invoice' from sale_order where upload_id = %s ", (self.id,))
           self.env.cr.execute("update a.state = 'sale' , a.invoice_status = 'to invoice' from sale_order_line a inner join sale_order b on a.order_id = b.id where a.upload_id = %s ", (self.id,))
           self.state = 'post'

        else:
           sql = "update a.state = 'draft' , a.invoice_status ='no' from sale_order_line a inner join sale_order b on a.order_id = b.id where a.upload_id %s "
           self.env.cr.execute("update state = 'draft' , invoice_status = 'no' from sale_order where upload_id = %s ", (self.id,))
           self.env.cr.execute("update a.state = 'draft' , a.invoice_status = 'no' from sale_order_line a inner join sale_order b on a.order_id = b.id where a.upload_id = %s ", (self.id,))
           self.state = 'done'



    def action_cancel(self):
        if self.state != 'done':
            so = self.env['sale.order'].search(['upload_id','=',self.id])
            so.unlink()
            upload_temp = self.env['sellout.upload.detail'].search(['upload_id','=',self.id])
            upload_temp.unlink()
            upload = self.env['sellout.upload'].search(['upload_id','=',self.id])
            upload.unlink()


            

class TempUploadCustomer(models.Model):
    _name = 'sellout.upload.detail.customer'
    _auto = False 

    upload_id   = fields.Many2one('sellout.upload.detail',string='upload id')
    company_id = fields.Many2one(
        comodel_name='res.company',
        index=True,
        default=lambda self: self.env.company)
    kode_ctm    = fields.Char(string='Kode Pelanggan')
    nama_ctm    = fields.Char(string='Nama Pelanggan')
    alamat      = fields.Char(string='Alamat')
    kelom_cust  = fields.Char(string='Kelompok Cust')
    kode_wilay  = fields.Char(string='Kode Wilayah')
    nama_wilay  = fields.Char(string='Nama Wilayah')
    customer_id = fields.Many2one(comodel_name='res.partner', string='Customer' )


    def init(self):
        tools.drop_view_if_exists(self._cr,'sellout_upload_detail_customer')
        self._cr.execute("""
            CREATE OR REPLACE VIEW sellout_upload_detail_customer AS ( 
                SELECT row_number() OVER () AS id, a.upload_id , a.company_id, a.kode_ctm, a.nama_ctm , a.alamat , a.kelom_cust , a.kode_wilay , a.nama_wilay, b.customer_id
                            FROM sellout_upload_detail a left outer join sellout_link_customer b 
                on a.kode_ctm = b.code and a.company_id = b.company_id
                            GROUP BY a.upload_id , a.company_id , a.kode_ctm,a.nama_ctm , a.alamat , a.kelom_cust , a.kode_wilay , a.nama_wilay, b.customer_id
                            ORDER BY a.upload_id , a.company_id , a.kode_ctm             
            )
        """)

    def _compute_customer(self):
        customer = self.env["link.customer"].search([('company_id','=',self.company_id.id),('kode_ctm','=',self.kode_ctm),('nama_ctm','=',self.nama_ctm)])
        if customer:
            customer.customer_id = self.customer.id 
        else:
            self.env["link.customer"].create({
                "company_id" : self.company_id.id,
                "kode_ctm"   :self.kode_ctm,
                "nama_ctm"   :self.nama_ctm,
                "customer_id": self.customer_id.id })


class TempUploadWilayah(models.Model):
    _name = 'sellout.upload.detail.wilayah'
    _auto = False 

    upload_id   = fields.Many2one('sellout.upload.detail',string='upload id')
    company_id = fields.Many2one(
        comodel_name='res.company',
        index=True,
        default=lambda self: self.env.company)
    kode_wilay   = fields.Char(string='Kode Wilayah')
    nama_wilay   = fields.Char(string='Nama Wilayah')
    district_id  = fields.Many2one(comodel_name='res.district', string='Wilayah')

    def init(self):
        tools.drop_view_if_exists(self._cr,'sellout_upload_detail_wilayah')
        self._cr.execute("""
            CREATE OR REPLACE VIEW sellout_upload_detail_wilayah AS ( 
                SELECT row_number() OVER () AS id, a.upload_id , a.company_id, a.kode_wilay, a.nama_wilay , b.district_id
                            FROM sellout_upload_detail a left outer join sellout_link_wilayah b 
                on a.company_id = b.company_id and a.kode_wilay = b.code
                            GROUP BY a.upload_id , a.company_id , a.kode_wilay, a.nama_wilay, b.district_id 
                            ORDER BY a.upload_id , a.company_id , a.kode_wilay            
            )
        """)
   

           

class TempUploadSalesman(models.Model):
    _name = 'sellout.upload.detail.salesman'
    _auto = False 

    upload_id   = fields.Many2one('sellout.upload.detail',string='upload id')
    company_id  = fields.Many2one(
        comodel_name='res.company',
        index=True,
        default=lambda self: self.env.company)
    kode_sls    = fields.Char(string='Kode Salesman')
    nama_sls    = fields.Char(string='Nama Salesman')
    user_id     = fields.Many2one(comodel_name='res.users', string='Saleperson')
    salesman_id = fields.Many2one(comodel_name='sellout.salesman', string='Salesman')

    def init(self):
        tools.drop_view_if_exists(self._cr,'sellout_upload_detail_salesman')
        self._cr.execute("""
            CREATE OR REPLACE VIEW sellout_upload_detail_salesman AS ( 
             SELECT row_number() OVER () AS id, a.upload_id , a.company_id, a.kode_sls, a.nama_sls , b.salesman_id
             FROM sellout_upload_detail a left outer join sellout_link_salesman b 
             on a.company_id = b.company_id and a.kode_sls = b.code
             GROUP BY a.upload_id , a.company_id , a.kode_sls, a.nama_sls, b.salesman_id
             ORDER BY a.upload_id , a.company_id , a.kode_sls
            )
        """)       

class TempUploadBarang(models.Model):
    _name = 'sellout.upload.detail.barang'
    _auto = False 

    upload_id   = fields.Many2one('sellout.upload.detail',string='upload id')
    company     = fields.Many2one(
        comodel_name='res.company',
        index=True,
        default=lambda self: self.env.company)
    kode_brg    = fields.Char(string='Kode Barang')
    nama_brg    = fields.Char(string='Nama Barang')
    barang_id   = fields.Many2one(comodel_name='product.product', string='Link Barang')

    def init(self):
        tools.drop_view_if_exists(self._cr,'sellout_upload_detail_barang')
        self._cr.execute("""
            CREATE OR REPLACE VIEW sellout_upload_detail_barang AS ( 
             SELECT row_number() OVER () AS id, a.upload_id,a.company_id as company, a.kode_brg, a.nama_brg, b.barang_id 
             FROM sellout_upload_detail a 
                 LEFT OUTER JOIN sellout_link_barang b ON a.company_id = b.company_id and a.kode_brg = b.code 
             GROUP BY a.upload_id,a.company_id,a.kode_brg,a.nama_brg,b.barang_id
             ORDER BY a.upload_id,a.company_id,a.kode_brg,a.nama_brg,b.barang_id 
            )
        """)      

    def write(self, values):
        print(" Write Method call ") 
        result = super(TempUploadBarang,self).write(values)
        return result


    